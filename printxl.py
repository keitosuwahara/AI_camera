import openpyxl as xl
import sqlite3
import datetime
import os
import shutil
from openpyxl.styles.borders import Border, Side

#アイデア
    #西暦と月日+名前で以前の出席を残す
    #月ごとのフォルダを自動で作り一年以上たったフォルダから消していく
    #ユーザー情報を最初に入力する時エクセルから入力させる→一気にdbに取り込む

today = datetime.date.today()#今日の日付
date = datetime.datetime.now()
year = date.year
month = date.month
day = date.day
hour = date.hour
minutes = date.minute
sec = date.second

#db内の情報をエクセルに転記するプログラム
def printXl(db):
    dbname =f"./database/{db}.db"
    conn = sqlite3.connect(dbname)
    #SQLiteを操作するためのカーソル,コントローラー
    cur = conn.cursor()
    #データベース内の表示
    info = []#仮でデータを入れておく##二次元配列にする→[[],[],[],[],[]]
    keys = []#キーを格納するリスト


    #選択したDBのそれぞれの行をリストで取り出す
    records = cur.execute(f"SELECT * FROM {db}")
    for record in records:
        info.append(list(record))
    #print(info)

    #キーの取得
    for desc in cur.description:
        keys.append(desc[0])

    conn.commit()

#↓エクセルの操作

    wb = xl.Workbook()
    ws = wb.worksheets[0]


    ws.cell(row = 1, column = 1, value = str(year)+"年")#西暦の転記
    ws.cell(row = 2, column = 1, value = str(month)+"月"+str(day)+"日")#月日の転記
    ws.cell(row = 3, column = 1, value = str(hour)+"時"+str(minutes)+"分")#時分の転記
    ws.cell(row = 4, column = 1, value = str(sec)+"秒時点")#秒の転記
    ws.cell(row = 5, column = 1, value = "DB名")
    ws.cell(row = 6, column = 1, value = db)#データベース名の転記
    

    
    
    tpm = []#datasに格納するために成形したリストを一時的に格納する
    datas = []#順番通りのリストを作る→[[DBの0番目のみ],[DBの1番目のみ],[DBの2番目のみ]]


    for i in range(0,len(info)):#リスト内リストの数
        for j in range(0,len(info[0])):#リスト内リストの要素数
            tpm.append(info[i][j])
    #print(tpm)[190721, '岩橋大地', 1, 0, 0, 200284, '桃崎奏斗', 1, 0, 0, 210103, '諏訪原慶斗', 1, 0, 1, 435755, '高田悠', 1, 1, 1, 454545, '志村', 1, 1, 1, 557855, '黒野怜奈', 1, 1, 1, 846556, 'トゴーフーバダムツェレン', 1, 1, 1]


    for iiii in range(0, len(info)-2):
        for iii in range(0, len(tpm), len(info[0])):
            datas.append(tpm[iii+iiii])#要素ごとに分けることには成功
    #print(datas)

    #datasを要素ごとの二次元配列にする関数(変換したいリスト, リスト内リスト数)
    def convert_1d_to_2d(datas, cols):
        return [datas[i:i + cols] for i in range(0, len(datas), cols)]

    stds = convert_1d_to_2d(datas, len(info))
    

    for index, key in enumerate(keys):#キーのみ転記する
        ws.cell(row=1, column=2+index, value=key)

    for ind, std in enumerate(stds):#キーに合った要素を転記する
        for i in range(0, len(std)):
            ws.cell(row=2+i, column=2+ind, value=std[i])
    
    os.makedirs(f"./database/xl_dir/{year}_{month}", exist_ok=True)
    wb.save(f'./database/xl_dir/{year}_{month}/{today}.xlsx')

    #2年前の出席シートを削除
    dir_path = './database/xl_dir'
    files = os.listdir(dir_path)  # ディレクトリ内のファイルリストを取得
    if len(files) >= 25:
        files.sort()  # ファイルリストを昇順に並び替え
        shutil.copytree(f'./database/xl_dir/{files[0]}', f'./static/ゴミ箱/xl_dirs/{files[0]}')#削除する前にゴミ箱に移す
        shutil.rmtree(f"./database/xl_dir/{files[0]}")# 先頭のファイル(=一番古いファイル名)を削除

    print("printxl関数が正常に実行されExcelに転記しました")

    #もう少し確実性のあるフォルダの日時の遡り方を考える

def createUserForm():#ユーザー情報を最初に入力する時エクセルから入力させる→一気にdbに取り込むプログラム
    wb = xl.Workbook()
    ws = wb.worksheets[0]
    wb.save(f'./登録用紙/{year}年度生徒情報登録用紙.xlsx')#ひとまず登録用紙を生成
    #例をあらかじめ載せておく
    #ふと枠の線を書いておきその中に入力させる

    #黒いふと枠線
    side1 = Side(style="thick", color='000000')
    border_aro = Border(top=side1, bottom=side1, left=side1, right=side1)

    for num in range(3,1000,1):#ふと枠を下まで書く
        ws[f'C{num}'].border = border_aro
        ws[f"D{num}"].border = border_aro
        ws[f"H{num}"].border = border_aro
        ws[f"I{num}"].border = border_aro

    # セルを結合
    ws.merge_cells("A4:B8")
    ws['A4'].value = "例文"
    #文字の大きさ
    ws['A4'].font = xl.styles.fonts.Font(size=40)

    # セルを結合
    ws.merge_cells("A10:B10")
    ws.merge_cells("A11:B11")
    ws.merge_cells("A12:B12")
    ws['A10'].value = "学籍番号は半角厳守"
    ws["A11"].value = "必ず太枠の中に入力してください"
    ws["A12"].value = "必ず一番上から入力してください"



    ws.cell(row = 2, column = 3, value = "学籍番号")#例の学籍番号
    ws.cell(row = 2, column = 4, value = "名前")#例の名前

    ws.cell(row = 3, column = 3, value = "257345")
    ws.cell(row = 4, column = 3, value = "235345")
    ws.cell(row = 5, column = 3, value = "285748")

    ws.cell(row = 3, column = 4, value = "山田孝之")
    ws.cell(row = 4, column = 4, value = "ウサイン・ボルト")
    ws.cell(row = 5, column = 4, value = "窪田正孝")

    # セルを結合
    ws.merge_cells("F4:G8")
    ws['F4'].value = "本番"
    #文字の大きさ
    ws['F4'].font = xl.styles.fonts.Font(size=40)

    ws.cell(row = 2, column = 8, value = "学籍番号")#本番の学籍番号
    ws.cell(row = 2, column = 9, value = "名前")#本番の名前

    wb.save(f'./登録用紙/{year}年度生徒情報登録用紙.xlsx')#登録用紙セーブ

def userFromXlToDB():
    #登録用紙xlを読み取る処理
    xl_path = f"./登録用紙/{year}年度生徒情報登録用紙.xlsx"
    userwb = xl.load_workbook(filename=xl_path)

    # シートのロード
    userws = userwb["Sheet"]

    #学籍番号と名前を入れておくリスト
    ids = []
    names = []


#家で読み込む場所を例文から本番に直しておく
    # 学籍番号セルの値取得 
    for num1 in range(3, 1000, 1):
        id_value = userws[f'C{num1}'].value
        if id_value == None:
            pass
        else:
            ids.append(id_value)
    #名前のセルの値取得
    for num2 in range(3, 1000, 1):
        name_value = userws[f'D{num2}'].value
        if name_value == None:
            pass
        else:
            names.append(name_value)


    

    # 取得した値の表示
    print('学籍番号', ids)
    print('名前', names)


    # ロードしたExcelファイルを閉じる
    userwb.close()
    #作成済みフォルダに入れる
    userwb.save(f"./登録用紙/作成済み生徒情報/{year}年度作成済み生徒情報.xlsx")
    print("userFromXlToDB()終了")


if __name__ == "__main__":
    createUserForm()
    userFromXlToDB()
    printXl("students")






